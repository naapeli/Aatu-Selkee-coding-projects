import tensorflow as tf
import numpy as np
from PIL import Image
import os
import matplotlib.pyplot as plt
from openpyxl import *


class Model:
    def __init__(self):
        self.nn_model = tf.keras.Sequential([
            tf.keras.layers.Dense(10, activation='relu', input_shape=(1024,)),
            tf.keras.layers.Dense(10, activation='relu'),
            tf.keras.layers.Dense(10, activation='softmax')
        ])
        self.nn_model.compile(optimizer=tf.keras.optimizers.Adam(0.001), loss='sparse_categorical_crossentropy', metrics=['accuracy'])
        self.nn_model.load_weights("nn_weights.h5")


    @staticmethod
    def save_train_data(number_of_each_digit):
        x_data = np.zeros((1, 1024))
        y_data = np.zeros((1, 1))
        directory = r"C:\Users\aatus\PycharmProjects\Python_projects_1\projects\Machine learning digit recognition software\data\Train"
        for digit_name in os.listdir(directory):
            i = 0
            for filename in os.listdir(os.path.join(directory, digit_name)):
                i += 1
                if i <= number_of_each_digit:
                    file = os.path.join(directory, digit_name, filename)
                    image = Image.open(file)
                    pixels = np.asarray(image)
                    pixels = pixels.reshape(1, 1024)
                    new_pixels = np.zeros((1, 1024))
                    for j, pixel in enumerate(pixels[0]):
                        if pixel >= 245:
                            new_pixels[0, j] = 1
                    x_data = np.vstack([x_data, new_pixels])
                    y_data = np.vstack([y_data, int(digit_name)])
        data = np.concatenate((x_data, y_data.reshape(-1, 1)), axis=1)
        data = data[np.random.permutation(data.shape[0])]
        return data[:, :-1], data[:, -1].reshape(-1, 1)

    def train(self, iterations, number_of_each_digit):
        x_data, y_data = self.save_train_data(number_of_each_digit)
        history = self.nn_model.fit(x_data[1:, :], y_data[1:, :], epochs=iterations, validation_split=0.2)
        self.nn_model.save_weights("nn_weights.h5")

        def plot_loss(history):
            plt.plot(history.history['loss'], label='loss')
            plt.plot(history.history['val_loss'], label='val_loss')
            plt.xlabel('Epoch')
            plt.ylabel('sparse_categorical_crossentropy')
            plt.legend()
            plt.grid(True)
            plt.show()

        def plot_accuracy(history):
            plt.plot(history.history['accuracy'], label='accuracy')
            plt.plot(history.history['val_accuracy'], label='val_accuracy')
            plt.xlabel('Epoch')
            plt.ylabel('Accuracy')
            plt.legend()
            plt.grid(True)
            plt.show()

        plt.figure(1)
        plot_loss(history)
        plt.figure(2)
        plot_accuracy(history)

    def predict(self, picture):
        predict_x = np.asarray(picture)[:, :, 0].reshape(1, 1024)

        wb = load_workbook(filename=r'C:\Users\aatus\PycharmProjects\Python_projects\projects\Machine learning digit recognition software\testing.xlsx')
        ws = wb.worksheets[0]
        i = 33
        for j, pixel in enumerate(predict_x[0]):
            j = j % 32
            if j == 31:
                i += 1
            if pixel >= 245:
                ws.cell(row=i+1, column=j+1).value = 1
            else:
                ws.cell(row=i+1, column=j+1).value = 0
        wb.save(r'C:\Users\aatus\PycharmProjects\Python_projects\projects\Machine learning digit recognition software\testing.xlsx')


        
        new_predict_x = np.zeros((1, 1024))
        for i, pixel in enumerate(predict_x[0]):
            if pixel >= 245:
                new_predict_x[0, i] = 1
            else:
                new_predict_x[0, i] = 0
        output = self.nn_model.predict(predict_x)
        number = np.argmax(output)
        return number

# malli = Model()
# malli.train(100, 2000)
