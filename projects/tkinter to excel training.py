import os
import openpyxl
import customtkinter as ctk
from tkinter import messagebox
from PIL import Image


class LoginFormWindow:
    def __init__(self, master):
        self.file_path = r"/projects/Python_projects/login_data.xlsx"
        self.master = master
        self.canvas = ctk.CTkFrame(self.master, fg_color="gray20")
        self.canvas.pack()
        window = ctk.CTkFrame(self.canvas, fg_color="gray20")
        window.grid(row=1, column=0)
        master.title("Login system")
        title = ctk.CTkLabel(self.canvas, text="Login", font=('Arial', 30), fg_color="gray20", width=500)
        title.grid(row=0, column=0)
        username_label = ctk.CTkLabel(window, text="Username", font=('Arial', 22))
        username_label.grid(row=0, column=0, pady=(20, 1))
        self.username = ctk.CTkEntry(window, width=500, height=35)
        self.username.grid(row=1, column=0, pady=5)
        password_label = ctk.CTkLabel(window, text="Password", font=('Arial', 22))
        password_label.grid(row=2, column=0, pady=(20, 1))
        self.password = ctk.CTkEntry(window, width=500, height=35)
        self.password.configure(show="*")
        self.password.grid(row=3, column=0, pady=(5, 1))
        bottom_frame = ctk.CTkFrame(window, fg_color="gray20")
        bottom_frame.grid(row=4, column=0)
        new_user_button = ctk.CTkButton(bottom_frame, text="New user", width=150, height=20, command=self.__new_user)
        new_user_button.grid(row=0, column=1)
        filler = ctk.CTkLabel(bottom_frame, width=350, height=25, text="")
        filler.grid(row=0, column=0, pady=(20, 3))
        login_button = ctk.CTkButton(window, text="Login", width=500, height=35, command=self.__login)
        login_button.grid(row=5, column=0)

    def __login(self):
        workbook = openpyxl.load_workbook(self.file_path)
        sheet = workbook.active
        if self.username.get() == "":
            messagebox.showwarning("Warning", "Insert your username!")
        else:
            if self.__username_exists(self.username.get(), sheet):
                if self.password.get() == self.__get_password(self.username.get(), sheet):
                    self.__clear()
                    self.__run_app()
                else:
                    messagebox.showwarning("Warning", "Wrong password!")
            else:
                messagebox.showwarning("Warning", "Username doesn't exists!")

        workbook.close()

    def __new_user(self):
        self.__clear()
        self.canvas = ctk.CTkFrame(self.master, fg_color="gray20")
        self.canvas.pack()
        frame_1 = ctk.CTkFrame(self.canvas, fg_color="gray20")
        frame_1.grid(row=0, column=0, pady=20)
        main_label_1 = ctk.CTkLabel(frame_1, text="Personal information", font=('Arial', 22))
        main_label_1.grid(row=0, column=0)

        frame_2 = ctk.CTkFrame(self.canvas, fg_color="gray20")
        frame_2.grid(row=1, column=0)
        label_name = ctk.CTkLabel(frame_2, text="Name")
        label_name.grid(row=0, column=0)
        self.name = ctk.CTkEntry(frame_2)
        self.name.grid(row=1, column=0)
        label_nationality = ctk.CTkLabel(frame_2, text="Nationality")
        label_nationality.grid(row=0, column=1)
        self.nationality = ctk.CTkEntry(frame_2)
        self.nationality.grid(row=1, column=1, padx=10)
        label_age = ctk.CTkLabel(frame_2, text="Age")
        label_age.grid(row=2, column=0)
        self.age = ctk.CTkEntry(frame_2)
        self.age.grid(row=3, column=0, padx=10)
        label_gender = ctk.CTkLabel(frame_2, text="Gender")
        label_gender.grid(row=2, column=1)
        default_gender = ctk.StringVar(value="Male")
        self.gender = ctk.CTkComboBox(frame_2, values=["Male", "Female", "Other"], variable=default_gender)
        self.gender.grid(row=3, column=1, padx=10)

        frame_3 = ctk.CTkFrame(self.canvas, fg_color="gray20")
        frame_3.grid(row=3, column=0, pady=(20, 0))
        main_label_2 = ctk.CTkLabel(frame_3, text="New login details", font=('Arial', 22))
        main_label_2.grid(row=0, column=0)
        new_username_label = ctk.CTkLabel(frame_3, text="New username", font=('Arial', 22))
        new_username_label.grid(row=1, column=0, pady=(20, 1))
        self.new_username = ctk.CTkEntry(frame_3, width=500, height=35)
        self.new_username.grid(row=2, column=0, pady=5)
        new_password_label = ctk.CTkLabel(frame_3, text="New password", font=('Arial', 22))
        new_password_label.grid(row=3, column=0, pady=(20, 1))
        self.new_password = ctk.CTkEntry(frame_3, width=500, height=35)
        self.new_password.configure(show="*")
        self.new_password.grid(row=4, column=0, pady=(5, 1))
        new_password_label_2 = ctk.CTkLabel(frame_3, text="Repeat new password", font=('Arial', 22))
        new_password_label_2.grid(row=5, column=0, pady=(20, 1))
        self.new_password_2 = ctk.CTkEntry(frame_3, width=500, height=35)
        self.new_password_2.configure(show="*")
        self.new_password_2.grid(row=6, column=0, pady=(5, 1))
        submit_button = ctk.CTkButton(frame_3, text="Submit", font=('Arial', 22), command=self.__submit)
        submit_button.grid(row=7, column=0)

    def __submit(self):
        new_password = self.new_password.get()
        new_password_2 = self.new_password_2.get()
        new_username = self.new_username.get()
        if not os.path.exists(self.file_path):
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            heading = ["Username", "Password", "Name", "Age", "Nationality", "Gender"]
            sheet.append(heading)
            workbook.save(self.file_path)
        if new_password == new_password_2:
            if self.name.get() != "" and self.age.get() != "" and self.gender.get() and self.nationality.get() != "":
                workbook = openpyxl.load_workbook(self.file_path)
                sheet = workbook.active
                if not self.__username_exists(new_username, sheet):
                    sheet.append([new_username, new_password, self.name.get(), self.age.get(), self.nationality.get(),
                                  self.gender.get()])
                    workbook.save(self.file_path)
                    self.__clear()
                    self.__init__(self.master)
                else:
                    workbook.close()
                    messagebox.showwarning("Warning", "Username already exists!")
            else:
                messagebox.showwarning("Warning", "All areas must be filled!")
        else:
            messagebox.showwarning("Warning", "Passwords don't match!")

    def __clear(self):
        self.canvas.destroy()

    @staticmethod
    def __username_exists(username, sheet) -> bool:
        for row in range(2, sheet.max_row + 1):
            if username == sheet.cell(row=row, column=1).value:
                return True
        return False

    @staticmethod
    def __get_password(username, sheet):
        for row in range(2, sheet.max_row + 1):
            if username == sheet.cell(row=row, column=1).value:
                return sheet.cell(row=row, column=2).value
        return "No password"

    def __run_app(self):
        self.canvas = ctk.CTkFrame(self.master, fg_color="gray20")
        self.canvas.pack()
        app = ctk.CTkLabel(self.canvas, text="You logged into my app.", font=('Arial', 22), width=500, height=500)
        app.pack()


root = ctk.CTk()
LoginFormWindow(root)
root.mainloop()
