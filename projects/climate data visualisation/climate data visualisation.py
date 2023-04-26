import tkinter as tk
import math
import openpyxl


months = ["January", "February", "March", "April", "May", "June", "July", "August", "September", "October", "November",
          "December"]
one_radius = 300
zero_radius = 150
width = 800
height = width


def create_circle(x, y, r, canvas):
    x0 = x - r
    y0 = x - r
    x1 = x + r
    y1 = x + r
    canvas.create_oval(x0, y0, x1, y1, width=1, outline="white")


class ClimateVisualisation:
    def __init__(self, master):
        master.title("Climate data")
        self.window = tk.Frame(master, bg="black")
        self.window.pack()
        label = tk.Label(self.window, text="Temperature each mont from 1888 to 2023", font=('Arial', 22), bg="black",
                         fg="white")
        label.pack(pady=10)
        self.canvas = tk.Canvas(self.window, width=width, height=height, bg="black")
        create_circle(width/2, height/2, one_radius, self.canvas)
        create_circle(width/2, height/2, zero_radius, self.canvas)
        self.canvas.pack()
        for i, month in enumerate(months):
            angle = 2 * math.pi * i / len(months) - math.pi / 2
            degrees = - angle * 180 / math.pi - 90
            r = one_radius + 3 * (height - 2 * one_radius) / 8
            x = r * math.cos(angle) + width / 2
            y = r * math.sin(angle) + height / 2
            self.canvas.create_text(x, y, angle=degrees, text=month, fill="white", font=('Arial', 12), anchor="center")
        self.canvas.pack()
        run_button = tk.Button(self.window, command=self.run, text="Run", width=40, height=5, bg="green")
        run_button.pack()
        self.first_cell = True
        self.previous_x = 0
        self.previous_y = 0
        self.year_text_id = self.canvas.create_text(width / 2, height / 2, text="1880", font=('Arial', 16), fill="white")
        self.current_row = 4
        self.current_cell = 1

    def run(self):
        file_path = r"C:\Users\aatus\PycharmProjects\pythonProject\projects\Python_projects\projects\climate data " \
                    r"visualisation\climate data.xlsx"
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        self.draw(self.current_cell, sheet, self.current_row)
        if self.current_row == sheet.max_row and self.current_cell == 13:
            return
        elif self.current_cell == 13:
            self.current_cell = 1
            self.current_row += 1
        else:
            self.current_cell += 1
        self.canvas.after(1, self.run)

    def draw(self, cell, sheet, row):
        if cell == 1:
            self.canvas.delete(self.year_text_id)
            year = sheet.cell(row=row, column=cell).value
            self.year_text_id = self.canvas.create_text(width / 2, height / 2, text=str(year), font=('Arial', 16),
                                                        fill="white")
        else:
            temperature = sheet.cell(row=row, column=cell).value
            angle = 2 * math.pi * (cell - 1) / len(months) - math.pi / 2
            if temperature != "***":
                r = (one_radius - zero_radius) * float(temperature) + zero_radius
                x = r * math.cos(angle) + width / 2
                y = r * math.sin(angle) + height / 2
                if self.first_cell:
                    self.first_cell = False
                    self.previous_x = x
                    self.previous_y = y
                    return
                self.canvas.create_line(self.previous_x, self.previous_y, x, y, fill=self.color_from_radius(r))
                self.previous_x = x
                self.previous_y = y

    def color_from_radius(self, radius: float):
        if radius >= zero_radius:
            green_blue = max(min(round(-200 * radius / (one_radius - zero_radius) +
                                       (255 * one_radius - 55 * zero_radius) / (one_radius - zero_radius)), 255), 0)
            rgb = (255, green_blue, green_blue)
            print(rgb)
            return self.hex_from_rgb(self, rgb)
        else:
            red_green = max(min(round(255 * radius / zero_radius), 255), 0)
            rgb = (red_green, red_green, 255)
            print(rgb)
            return self.hex_from_rgb(self, rgb)

    @staticmethod
    def hex_from_rgb(self, rgb):
        """translates a rgb tuple of int to a tkinter friendly color code
        """
        return "#%02x%02x%02x" % rgb


if __name__ == "__main__":
    root = tk.Tk()
    window = ClimateVisualisation(root)
    root.mainloop()
