from PIL import Image
import numpy as np
import tkinter as tk
from openpyxl import *
import pandas as pd



# load the image
image = Image.open(r'C:\Users\aatus\PycharmProjects\Python_projects_1\projects\Machine learning digit recognition software\data\Train\5\500.jpg')
image2 = Image.open(r'C:\Users\aatus\OneDrive\Documents\GitHub\Python_projects\projects\Machine learning digit recognition software\testing_picture.jpg')
# convert image to numpy array
data = np.asarray(image)

wb = load_workbook(filename=r'C:\Users\aatus\PycharmProjects\Python_projects\projects\Machine learning digit recognition software\testing.xlsx')
ws = wb.worksheets[0]

data = data.reshape(1, 1024)
i = 0
for j, pixel in enumerate(data[0]):
	j = j % 32
	if j == 31:
		i += 1
	if pixel >= 245:
		ws.cell(row=i+1, column=j+1).value = 1
	else:
		ws.cell(row=i+1, column=j+1).value = 0

wb.save(r'C:\Users\aatus\PycharmProjects\Python_projects\projects\Machine learning digit recognition software\testing.xlsx')

'''
pixels = np.asarray(image)
pixels = pixels.reshape(1, 1024)
new_pixels = np.zeros((1, 1024))
for j, pixel in enumerate(pixels[0]):
    if pixel >= 10:
        new_pixels[0, j] = 255
'''
